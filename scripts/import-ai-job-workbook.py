#!/usr/bin/env python3
"""Import AI_Job_Opportunity_Forecasting_Dashboard workbook into growth + recommendations JSON."""

from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "AI_Job_Opportunity_Forecasting_Dashboard_India(1).xlsx"
DATA_DIR = ROOT / "data"
PUBLIC_DIR = ROOT / "dashboard" / "public"

GEO_STATE_ALIASES = {
    "Delhi-NCR": "Delhi",
    "Delhi NCR": "Delhi",
    "NCR": "Delhi",
    "Orissa": "Odisha",
    "U.P.": "Uttar Pradesh",
    "UP": "Uttar Pradesh",
}


def slug(text: str) -> str:
    return re.sub(r"-+$", "", re.sub(r"^-+", "", re.sub(r"[^a-z0-9]+", "-", text.lower().replace("&", "and"))))


def parse_num(raw) -> float:
    if raw is None or raw == "":
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    cleaned = re.sub(r"[^0-9.\-]", "", str(raw))
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def clean_header(h: str) -> str:
    return re.sub(r"\s+", " ", str(h or "").strip())


def get_val(row: dict, *needles: str, default=""):
    """Match column by substring(s) — headers are normalized (newlines → spaces)."""
    if not row:
        return default
    needles_l = [n.lower() for n in needles]
    for key, val in row.items():
        k = str(key).lower()
        if all(n in k for n in needles_l):
            return val if val is not None else default
    for key, val in row.items():
        k = str(key).lower()
        if any(n in k for n in needles_l):
            return val if val is not None else default
    return default


def sheet_rows(wb, name: str) -> list[dict]:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = 0
    for i, row in enumerate(rows[:25]):
        if sum(1 for c in row if c not in (None, "")) >= 3:
            header_idx = i
            break
    headers = [clean_header(c) for c in rows[header_idx]]
    out: list[dict] = []
    for row in rows[header_idx + 1 :]:
        if all(c in (None, "") for c in row):
            continue
        obj = {}
        for h, v in zip(headers, row):
            if not h:
                continue
            if v is None:
                obj[h] = ""
            elif isinstance(v, float) and v == int(v):
                obj[h] = int(v)
            else:
                obj[h] = v
        if any(str(v).strip() for v in obj.values()):
            out.append(obj)
    return out


def normalize_state(raw: str) -> str:
    s = str(raw or "").strip()
    if not s or s.lower() in {"multi-state", "pan-india", "multi-state (tn, up, karnataka, gujarat)"}:
        return ""
    if "," in s:
        s = s.split(",")[0].strip()
    return GEO_STATE_ALIASES.get(s, s)


def confidence_label(raw: str) -> str:
    v = str(raw or "").strip().upper()
    if v in {"H", "HIGH"}:
        return "High"
    if v in {"M", "MEDIUM", "MED"}:
        return "Medium"
    if v in {"L", "LOW"}:
        return "Low"
    return str(raw or "Medium")


def priority_from_gap(severity: str, gap_pct: float) -> str:
    s = str(severity or "").lower()
    if "critical" in s or gap_pct >= 35:
        return "Critical"
    if "high" in s or gap_pct >= 25:
        return "High"
    if gap_pct >= 15:
        return "Medium"
    return "Low"


def pipeline_to_main_dataset(pipeline: list[dict]) -> list[dict]:
    rows = []
    for p in pipeline:
        state = normalize_state(p.get("State", ""))
        city = str(p.get("City/District", "")).strip()
        location = ", ".join(x for x in [city, state] if x)
        direct = parse_num(get_val(p, "direct jobs", "announced"))
        indirect = parse_num(get_val(p, "indirect jobs"))
        total = parse_num(get_val(p, "total jobs")) or (direct + indirect)
        announced = get_val(p, "announced", "approved date")
        rows.append(
            {
                "Region": state or "India",
                "State": state or "India",
                "Department / Industry": str(p.get("Sector", "")).strip(),
                "Sector": str(p.get("Sector", "")).strip(),
                "Sub-Sector": str(p.get("Sector", "")).strip(),
                "Investment Project / Initiative": str(p.get("Company / Project Name", "")).strip(),
                "Investment Value (INR Cr)": parse_num(get_val(p, "investment", "crore")),
                "Skill Type": "Skilled",
                "Job Category": str(p.get("Project Stage", "")).strip(),
                "Projected Vacancies": int(total or direct),
                "Direct Jobs": int(direct),
                "Indirect Jobs": int(indirect),
                "Start Date": str(get_val(p, "expected hiring", "start")).strip(),
                "Expected Completion": "",
                "Hiring Period": str(get_val(p, "expected hiring", "start")).strip(),
                "Location": location,
                "City/District": city,
                "Key Skills Required": str(p.get("Project Stage", "")).strip(),
                "Source / Reference": str(p.get("Source", "")).strip(),
                "Confidence Level": confidence_label(get_val(p, "confidence")),
                "Additional Insights": f"Stage: {p.get('Project Stage', '')}. Announced: {announced}.",
                "Project_ID": str(p.get("Project_ID", "")).strip(),
                "Project Stage": str(p.get("Project Stage", "")).strip(),
            }
        )
    return rows


def build_sectors(main_dataset: list[dict], forecast: list[dict]) -> list[dict]:
    by_sector: dict[str, dict] = {}
    forecast_map = {}
    for f in forecast:
        key = (str(f.get("Sector", "")).strip(), normalize_state(f.get("State", "")))
        forecast_map[key] = f

    for row in main_dataset:
        name = str(row.get("Sector", "")).strip() or "General"
        sector = by_sector.setdefault(
            name,
            {
                "id": slug(name),
                "name": name,
                "slug": slug(name),
                "policy": "",
                "investmentSignal": "high",
                "investmentScore": 75,
                "growthRate": 20,
                "growthMultiplier": 1.2,
                "liveOnSite": True,
                "sourceUrl": "",
                "baseline": {"listings": 0, "vacancies": 0},
                "predictedOpenings6m": 0,
                "predictedOpenings12m": 0,
                "confidence": 80,
                "timeline": [],
                "typicalRoles": [],
                "educationDemand": {"skilled": 100},
                "districtHotspots": [],
                "aiRationale": "",
                "keywords": [name],
                "investmentCr": 0,
                "projectCount": 0,
                "projects": [],
            },
        )
        jobs = parse_num(row.get("Projected Vacancies"))
        inv = parse_num(row.get("Investment Value (INR Cr)"))
        sector["predictedOpenings12m"] += int(jobs)
        sector["investmentCr"] += inv
        sector["projectCount"] += 1
        state = str(row.get("State", "")).strip()
        city = str(row.get("City/District", "")).strip()
        if city and city not in sector["districtHotspots"]:
            sector["districtHotspots"].append(city)
        elif state and state not in sector["districtHotspots"]:
            sector["districtHotspots"].append(state)
        sector["projects"].append(
            {
                "name": str(row.get("Investment Project / Initiative", "")),
                "subSector": str(row.get("Sub-Sector", name)),
                "investmentCr": inv,
                "vacancies": int(jobs),
                "startDate": str(row.get("Start Date", "")),
                "expectedCompletion": "",
                "hiringPeriod": str(row.get("Hiring Period", "")),
                "location": str(row.get("Location", "")),
                "skillType": str(row.get("Skill Type", "")),
                "jobCategory": str(row.get("Job Category", "")),
                "keySkillsRequired": str(row.get("Key Skills Required", "")),
                "sourceReference": str(row.get("Source / Reference", "")),
                "confidenceLevel": str(row.get("Confidence Level", "")),
                "additionalInsights": str(row.get("Additional Insights", "")),
            }
        )

    for sector in by_sector.values():
        sector["predictedOpenings6m"] = int(sector["predictedOpenings12m"] * 0.45)
        sector["aiRationale"] = (
            f"{sector['name']} — ₹{sector['investmentCr']:,.0f} Cr tracked investment, "
            f"{sector['predictedOpenings12m']:,} projected jobs across {sector['projectCount']} pipeline project(s)."
        )
        if sector["investmentCr"] >= 50000:
            sector["investmentSignal"] = "high"
            sector["investmentScore"] = 90
        elif sector["investmentCr"] >= 10000:
            sector["investmentSignal"] = "medium"
            sector["investmentScore"] = 70
        else:
            sector["investmentSignal"] = "low"
            sector["investmentScore"] = 55

    return sorted(by_sector.values(), key=lambda s: s["predictedOpenings12m"], reverse=True)


def build_executive_summary(vacancy_gap: list[dict], main_dataset: list[dict]) -> list[dict]:
    agg: dict[tuple[str, str], dict] = {}
    for row in vacancy_gap:
        state = normalize_state(row.get("State", ""))
        sector = str(row.get("Sector", "")).strip()
        key = (state or "India", sector)
        item = agg.setdefault(
            key,
            {
                "Region": state or "India",
                "Industry": sector,
                "Total Investment (INR Cr)": 0,
                "Projected Jobs (Total)": 0,
                "Major Locations": state,
                "Growth Outlook": str(row.get("Gap Severity", "")),
                "Investment Priority": "★★★★☆" if "high" in str(row.get("Gap Severity", "")).lower() else "★★★☆☆",
                "Key Drivers": str(row.get("Top Constraint", "")),
            },
        )
        item["Projected Jobs (Total)"] += int(parse_num(row.get("Total Projected Demand (12-mo)")))

    for row in main_dataset:
        state = normalize_state(row.get("State", ""))
        sector = str(row.get("Sector", "")).strip()
        key = (state or "India", sector)
        item = agg.setdefault(
            key,
            {
                "Region": state or "India",
                "Industry": sector,
                "Total Investment (INR Cr)": 0,
                "Projected Jobs (Total)": 0,
                "Major Locations": str(row.get("Location", state)),
                "Growth Outlook": "High ▲",
                "Investment Priority": "★★★★☆",
                "Key Drivers": str(row.get("Source / Reference", "")),
            },
        )
        item["Total Investment (INR Cr)"] += parse_num(row.get("Investment Value (INR Cr)"))

    return list(agg.values())


def is_aggregate_label(label: str) -> bool:
    s = str(label or "").strip()
    if not s or len(s) > 72:
        return True
    lower = s.lower()
    if lower.startswith("total"):
        return True
    if lower.startswith("national"):
        return True
    if lower.startswith("methodology"):
        return True
    if lower.startswith("source project"):
        return True
    return False


def normalize_vacancy_gap(rows: list[dict]) -> list[dict]:
    out = []
    for row in rows:
        sector = str(row.get("Sector", "")).strip()
        if is_aggregate_label(sector):
            continue
        out.append(
            {
                "Sector": sector,
                "State": normalize_state(row.get("State", "")),
                "Current Open Vacancies (est.)": int(parse_num(get_val(row, "current open", "vacancies"))),
                "New Vacancies from Pipeline (12-mo)": int(parse_num(get_val(row, "new vacancies", "pipeline"))),
                "Total Projected Demand (12-mo)": int(parse_num(get_val(row, "total projected", "demand"))),
                "Annual Skilled-Talent Supply": int(parse_num(get_val(row, "skilled-talent", "supply"))),
                "Net Vacancy Gap": int(parse_num(get_val(row, "net vacancy gap"))),
                "Gap as % of Demand": parse_num(get_val(row, "gap as %", "demand")),
                "Gap Severity": str(get_val(row, "gap severity")).strip(),
                "Avg. Time-to-Fill (days)": int(parse_num(get_val(row, "time-to-fill"))),
                "Median CTC (₹ lakh p.a.)": parse_num(get_val(row, "median ctc")),
                "YoY Demand Growth %": parse_num(get_val(row, "yoy demand")),
                "Top Constraint": str(get_val(row, "top constraint")).strip(),
                "Recommended Action": str(get_val(row, "recommended", "action")).strip(),
                "Confidence": confidence_label(get_val(row, "confidence")),
                "Source": str(get_val(row, "source")).strip(),
            }
        )
    return out


def normalize_forecast(rows: list[dict]) -> list[dict]:
    out = []
    for row in rows:
        sector = str(row.get("Sector", "")).strip()
        if is_aggregate_label(sector):
            continue
        out.append(
            {
                "Sector": sector,
                "State": normalize_state(row.get("State", "")),
                "Investment in Scope (₹ Cr)": parse_num(get_val(row, "investment in scope")),
                "Direct Jobs per ₹100cr": parse_num(get_val(row, "direct jobs", "100cr")),
                "Indirect Multiplier": parse_num(get_val(row, "indirect", "multiplier")),
                "3-Month Jobs": int(parse_num(get_val(row, "3-month jobs"))),
                "6-Month Jobs": int(parse_num(get_val(row, "6-month jobs"))),
                "12-Month Jobs": int(parse_num(get_val(row, "12-month jobs"))),
                "24-Month Jobs": int(parse_num(get_val(row, "24-month jobs"))),
                "3-Mo CI Low": int(parse_num(get_val(row, "3-mo ci low"))),
                "3-Mo CI High": int(parse_num(get_val(row, "3-mo ci high"))),
                "12-Mo CI Low": int(parse_num(get_val(row, "12-mo ci low"))),
                "12-Mo CI High": int(parse_num(get_val(row, "12-mo ci high"))),
                "Model Confidence": str(get_val(row, "model confidence")).strip(),
            }
        )
    return out


def is_heatmap_state_row(state: str) -> bool:
    s = str(state or "").strip()
    if not s or len(s) > 48:
        return False
    lower = s.lower()
    if lower.startswith("index construction"):
        return False
    if "recompute monthly" in lower or "darker red" in lower:
        return False
    return True


def normalize_heatmap(rows: list[dict]) -> list[dict]:
    if not rows:
        return []
    first = rows[0]
    sector_cols = [k for k in first.keys() if k != "State / Sector"]
    out = []
    for row in rows:
        state = str(row.get("State / Sector", "")).strip()
        if not is_heatmap_state_row(state):
            continue
        entry = {"State": state, "State / Sector": state}
        for col in sector_cols:
            entry[col] = parse_num(row.get(col))
        out.append(entry)
    return out


def parse_dashboard_sheet(wb) -> dict:
    """Parse the Dashboard sheet layout (KPI row, horizon table, top states)."""
    ws = wb["Dashboard"]
    rows = list(ws.iter_rows(values_only=True))
    kpis = {
        "pliJobsCumulative": 0,
        "fdiInflowUsdBn": 0,
        "startupFundingUsdBn": 0,
        "trackedPipelineInvestmentCr": 0,
        "forecastHorizons": [],
        "topStatesByIntensity": [],
    }

    in_top_states = False
    for i, row in enumerate(rows):
        cells = list(row)
        label = str(cells[0] or "").strip()

        # KPI value row: numbers in cols 0, 3, 6, 9 under header labels
        if (
            isinstance(cells[0], (int, float))
            and len(cells) > 3
            and isinstance(cells[3], (int, float))
        ):
            prev = str(rows[i - 1][0] or "") if i > 0 else ""
            if "PLI Jobs" in prev or "Total PLI" in prev:
                kpis["pliJobsCumulative"] = int(parse_num(cells[0]))
                kpis["fdiInflowUsdBn"] = parse_num(cells[3])
                kpis["startupFundingUsdBn"] = parse_num(cells[6]) if len(cells) > 6 else 0
                kpis["trackedPipelineInvestmentCr"] = parse_num(cells[9]) if len(cells) > 9 else 0

        # National forecast horizons
        if label in ("3 Months", "6 Months", "12 Months", "24 Months"):
            kpis["forecastHorizons"].append(
                {
                    "horizon": label,
                    "jobs": int(parse_num(cells[1] if len(cells) > 1 else 0)),
                    "low": int(parse_num(cells[2] if len(cells) > 2 else 0)),
                    "high": int(parse_num(cells[3] if len(cells) > 3 else 0)),
                }
            )

        if label == "Rank" and str(cells[1] or "").strip() == "State":
            in_top_states = True
            continue

        if in_top_states:
            if not isinstance(cells[0], (int, float)) or not cells[1]:
                in_top_states = False
            else:
                kpis["topStatesByIntensity"].append(
                    {
                        "rank": int(parse_num(cells[0])),
                        "state": str(cells[1] or "").strip(),
                        "intensity": parse_num(cells[2] if len(cells) > 2 else 0),
                    }
                )

    return kpis


def build_growth_report(wb) -> dict:
    pipeline = sheet_rows(wb, "Project_Pipeline")
    vacancy_gap = normalize_vacancy_gap(sheet_rows(wb, "Vacancy_Gap_Analysis"))
    forecast = normalize_forecast(sheet_rows(wb, "Job_Forecast_3_6_12_24M"))
    main_dataset = pipeline_to_main_dataset(pipeline)
    sectors = build_sectors(main_dataset, forecast)
    dashboard_kpis = parse_dashboard_sheet(wb)

    total_jobs_12m = sum(int(parse_num(r.get("12-Month Jobs"))) for r in forecast) or sum(
        int(parse_num(r.get("Projected Vacancies"))) for r in main_dataset
    )
    total_investment = sum(parse_num(r.get("Investment Value (INR Cr)")) for r in main_dataset)
    states = {normalize_state(r.get("State", "")) for r in main_dataset if normalize_state(r.get("State", ""))}

    sheets = {
        "mainDataset": main_dataset,
        "projectPipeline": main_dataset,
        "projectPipelineSheet": pipeline,
        "controlPanel": sheet_rows(wb, "Control_Panel"),
        "vacancyGapAnalysis": vacancy_gap,
        "jobForecast": forecast,
        "sectorSubSectorMaster": sheet_rows(wb, "Sector_SubSector_Master"),
        "stateCityMaster": sheet_rows(wb, "State_City_Master"),
        "hiringHeatmap": normalize_heatmap(sheet_rows(wb, "Hiring_Heatmap_Matrix")),
        "skillDemandForecast": sheet_rows(wb, "Skills_Roles_Demand"),
        "topOpportunities": sheet_rows(wb, "Top_Employers_Hiring"),
        "employmentRanking": forecast,
        "executiveSummary": build_executive_summary(vacancy_gap, main_dataset),
        "investmentToEmployment": sheet_rows(wb, "Investment_to_Employment"),
        "aiInsights": sheet_rows(wb, "AI_Insights_Log"),
        "alertsWatchlist": sheet_rows(wb, "Alerts_Watchlist"),
        "fdiTracker": sheet_rows(wb, "FDI_Inflow_Tracker"),
        "pliTracker": sheet_rows(wb, "PLI_Scheme_Tracker"),
        "startupTracker": sheet_rows(wb, "Startup_Funding_Tracker"),
        "historicalTrend": sheet_rows(wb, "Historical_Trend_Comparison"),
        "economicIndicators": sheet_rows(wb, "Economic_Indicators"),
        "dataSources": sheet_rows(wb, "Data_Sources"),
        "methodologyAssumptions": sheet_rows(wb, "Methodology_Assumptions"),
        "dashboardKpis": dashboard_kpis,
    }

    return {
        "generatedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "model": "ai-job-opportunity-forecasting-v1",
        "modelNote": "Imported from AI Job Opportunity Forecasting Dashboard workbook (PIB, DPIIT, Naukri JobSpeak, Tracxn, India Investment Grid, NSDC).",
        "stateCode": "IN",
        "summary": {
            "sectorCount": len(sectors),
            "totalPredicted6m": int(total_jobs_12m * 0.45),
            "totalPredicted12m": int(total_jobs_12m),
            "highGrowthSectors": sum(1 for s in sectors if s["investmentSignal"] == "high"),
            "avgConfidence": 85,
            "totalInvestmentCr": int(total_investment),
            "projectCount": len(main_dataset),
            "districtCount": len(states),
            "topSectors": [
                {
                    "id": s["id"],
                    "name": s["name"],
                    "predicted12m": s["predictedOpenings12m"],
                    "confidence": s["confidence"],
                }
                for s in sectors[:5]
            ],
        },
        "sectors": sectors,
        "workbook": {
            "source": "AI_Job_Opportunity_Forecasting_Dashboard_India(1).xlsx",
            "sheets": sheets,
        },
        "meta": {
            "lastSync": {
                "synced_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
                "sector_count": len(sectors),
                "status": "workbook",
                "message": "Loaded all sheets from AI Job Opportunity Forecasting Dashboard",
            },
            "sectorCount": len(sectors),
        },
    }


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        print(f"Workbook not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    growth = build_growth_report(wb)

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    PUBLIC_DIR.mkdir(parents=True, exist_ok=True)

    growth_path = DATA_DIR / "upGrowthInvestmentReport.json"
    growth_public = PUBLIC_DIR / "upGrowthInvestmentReport.json"
    for p in (growth_path, growth_public):
        with open(p, "w", encoding="utf-8") as f:
            json.dump(growth, f, indent=2, ensure_ascii=False)
        print(f"Wrote {p}")

    print("Run: node scripts/build-recommendations-from-growth.js")


if __name__ == "__main__":
    main()
